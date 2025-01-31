import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def highlight_differences(row, base_columns):
    """Highlight differences in the row for the specified base columns."""
    styles = []
    for col in base_columns:
        col_file1 = f"{col}_file1"
        col_file2 = f"{col}_file2"
        if col_file1 in row.index and col_file2 in row.index:
            if row[col_file1] != row[col_file2]:
                styles.append('background-color: yellow')
                styles.append('background-color: yellow')
            else:
                styles.append('')
                styles.append('')
        else:
            styles.append('')
            styles.append('')
    return styles

def save_styled_excel(dataframe, base_columns):
    """Save the styled DataFrame as an Excel file with highlights."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Differences"

    # Write header
    for col_num, column_title in enumerate(dataframe.columns, start=1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Write data and apply styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row_num, row in enumerate(dataframe.itertuples(index=False), start=2):
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            # Check if the cell is part of a difference and apply highlight
            col_name = dataframe.columns[col_num - 1]
            base_col = col_name.replace('_file1', '').replace('_file2', '')
            if base_col in base_columns and '_file' in col_name:
                other_col_name = col_name.replace('_file1', '_file2') if '_file1' in col_name else col_name.replace('_file2', '_file1')
                if other_col_name in dataframe.columns:
                    other_value = row[dataframe.columns.get_loc(other_col_name)]
                    if value != other_value:
                        cell.fill = yellow_fill

    with BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

def main():
    st.title("CSV File Comparison Tool")

    # Upload the first and second files
    file1 = st.file_uploader("Upload the first CSV file", type="csv", key="file1")
    file2 = st.file_uploader("Upload the second CSV file", type="csv", key="file2")

    if file1 and file2:
        # Read the uploaded CSV files
        df1 = pd.read_csv(file1)
        df2 = pd.read_csv(file2)

        # Check if the structure of the two files is the same
        if df1.columns.tolist() != df2.columns.tolist():
            st.error("The two files have different column structures. Please upload files with matching columns.")
            return

        # Align both DataFrames by index and columns
        df1 = df1.sort_index().reset_index(drop=True)
        df2 = df2.sort_index().reset_index(drop=True)

        # Add suffixes to differentiate file1 and file2 data
        df1 = df1.add_suffix('_file1')
        df2 = df2.add_suffix('_file2')

        # Combine the two DataFrames for comparison
        combined = pd.concat([df1, df2], axis=1)

        # Identify columns without suffixes for comparison
        base_columns = [col.replace('_file1', '') for col in df1.columns if '_file1' in col]

        # Identify rows with differences
        diff_mask = (df1.values != df2.values)  # Compare values directly
        differing_rows = combined[np.any(diff_mask, axis=1)]

        if differing_rows.empty:
            st.success("The two files are identical!")
        else:
            st.warning("Differences found between the files.")

            # Highlight differences
            styled_diff = differing_rows.style.apply(lambda row: highlight_differences(row, base_columns), axis=1)

            # Display differing rows with highlights
            st.write("Rows with differences:")
            st.dataframe(differing_rows, use_container_width=True)

            # Add download functionality for differing rows
            filename1 = file1.name.split('.')[0]
            filename2 = file2.name.split('.')[0]
            output_filename = f"{filename1}_vs_{filename2}.xlsx"

            # Save the differing rows to an Excel file with highlights
            excel_data = save_styled_excel(differing_rows, base_columns)

            # Provide a download link
            st.download_button(
                label="Download Highlighted Differences (Excel)",
                data=excel_data,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    main()
